#!/usr/bin/env python3

import json
import pandas as pd
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import numpy as np
import os
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')


class EnhancedExcelReport:
    def __init__(self, json_file, visualization_data_file=None):
        """Initialize report generator with data files"""
        self.json_file = json_file
        self.visualization_data_file = visualization_data_file
        
        # Load benchmark data
        with open(json_file, 'r') as f:
            self.data = json.load(f)
        
        # Load visualization data if available
        self.viz_data = None
        if visualization_data_file and os.path.exists(visualization_data_file):
            with open(visualization_data_file, 'r') as f:
                self.viz_data = json.load(f)
            print(f"Loaded visualization data with {len(self.viz_data.get('timestamps', []))} data points")
    
    def create_time_series_plots(self, output_dir):
        """Create time series plots for all metrics"""
        if not self.viz_data or len(self.viz_data.get('timestamps', [])) == 0:
            print("No visualization data available for plotting")
            return {}
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Use a clean style
        plt.style.use('default')
        
        plots = {}
        
        # Common plot settings
        fig_size = (10, 6)
        colors = {
            'hz': '#1f77b4',
            'jitter': '#ff7f0e',
            'bandwidth': '#2ca02c',
            'throughput': '#d62728',
            'cpu': '#9467bd',
            'memory': '#8c564b',
            'temperature': '#e377c2'
        }
        
        # Helper function to create a plot
        def create_plot(data_key, title, ylabel, color, filename):
            if data_key not in self.viz_data or len(self.viz_data[data_key]) == 0:
                return None
                
            plt.figure(figsize=fig_size)
            
            # Get data
            timestamps = self.viz_data['timestamps'][:len(self.viz_data[data_key])]
            values = self.viz_data[data_key]
            
            # Main plot
            plt.plot(timestamps, values, color=color, linewidth=2, alpha=0.8)
            
            # Calculate statistics
            avg_val = np.mean(values)
            std_val = np.std(values)
            min_val = np.min(values)
            max_val = np.max(values)
            
            # Add average line
            plt.axhline(y=avg_val, color='red', linestyle='--', linewidth=1.5, 
                       label=f'Average: {avg_val:.2f}')
            
            # Add shaded area for standard deviation
            if std_val > 0:
                plt.fill_between(timestamps, avg_val - std_val, avg_val + std_val, 
                               alpha=0.2, color='gray', label=f'±1 STD: {std_val:.2f}')
            
            # Formatting
            plt.title(title, fontsize=16, fontweight='bold', pad=20)
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel(ylabel, fontsize=12)
            plt.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)
            plt.legend(loc='upper right', fontsize=10)
            
            # Add text box with statistics
            stats_text = f'Min: {min_val:.2f}\nMax: {max_val:.2f}\nAvg: {avg_val:.2f}\nSTD: {std_val:.2f}'
            plt.text(0.02, 0.98, stats_text, transform=plt.gca().transAxes, 
                    verticalalignment='top', fontsize=10,
                    bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
            
            # Tight layout
            plt.tight_layout()
            
            # Save
            plot_path = os.path.join(output_dir, filename)
            plt.savefig(plot_path, dpi=150, bbox_inches='tight', facecolor='white')
            plt.close()
            
            return plot_path
        
        # Create all plots
        plots['frequency'] = create_plot('hz', 'LiDAR Publishing Frequency Over Time', 
                                       'Frequency (Hz)', colors['hz'], 'frequency_plot.png')
        
        plots['jitter'] = create_plot('jitter', 'Message Jitter Over Time', 
                                    'Jitter (ms)', colors['jitter'], 'jitter_plot.png')
        
        plots['bandwidth'] = create_plot('bandwidth', 'Network Bandwidth Usage Over Time', 
                                       'Bandwidth (Mbps)', colors['bandwidth'], 'bandwidth_plot.png')
        
        plots['throughput'] = create_plot('throughput', 'Point Cloud Throughput Over Time', 
                                        'Throughput (K points/sec)', colors['throughput'], 'throughput_plot.png')
        
        plots['cpu'] = create_plot('cpu', 'CPU Usage Over Time', 
                                 'CPU Usage (%)', colors['cpu'], 'cpu_plot.png')
        
        plots['memory'] = create_plot('memory', 'Memory Usage Over Time', 
                                    'Memory Usage (%)', colors['memory'], 'memory_plot.png')
        
        plots['temperature'] = create_plot('temperature', 'CPU Temperature Over Time', 
                                         'Temperature (°C)', colors['temperature'], 'temperature_plot.png')
        
        # Remove None values
        plots = {k: v for k, v in plots.items() if v is not None}
        
        print(f"Created {len(plots)} plots")
        
        return plots
    
    def generate_excel_with_graphs(self, output_file):
        """Generate Excel report with embedded graphs"""
        print(f"Generating Excel report with graphs: {output_file}")
        
        # First, generate the basic Excel report
        self._generate_basic_excel(output_file)
        
        # Create plots
        plot_dir = os.path.join(os.path.dirname(output_file), 'temp_plots')
        plots = self.create_time_series_plots(plot_dir)
        
        if not plots:
            print("No plots generated, skipping graph integration")
            return
        
        # Load the workbook and add graphs
        print("Adding graphs to Excel...")
        wb = load_workbook(output_file)
        
        # Create a new sheet for graphs
        if 'Graphs' in wb.sheetnames:
            del wb['Graphs']
        
        ws_graphs = wb.create_sheet('Graphs', 1)  # Insert after Summary
        
        # Format the sheet
        ws_graphs.sheet_properties.tabColor = "0066CC"
        
        # Add title
        ws_graphs['A1'] = 'Performance Metrics Time Series Analysis'
        ws_graphs['A1'].font = Font(bold=True, size=18)
        ws_graphs['A1'].alignment = Alignment(horizontal='center')
        ws_graphs.merge_cells('A1:H1')
        
        # Add description
        ws_graphs['A3'] = 'The following graphs show the performance metrics over the entire benchmark duration.'
        ws_graphs['A3'].font = Font(italic=True, size=11)
        ws_graphs.merge_cells('A3:H3')
        
        # Define plot layout (2 columns)
        plot_positions = [
            ('frequency', 'A5', 'Publishing Frequency Analysis'),
            ('jitter', 'J5', 'Message Jitter Analysis'),
            ('bandwidth', 'A30', 'Network Bandwidth Analysis'),
            ('throughput', 'J30', 'Throughput Analysis'),
            ('cpu', 'A55', 'CPU Usage Analysis'),
            ('memory', 'J55', 'Memory Usage Analysis'),
            ('temperature', 'A80', 'Temperature Analysis')
        ]
        
        # Add each plot
        for key, cell, title in plot_positions:
            if key in plots and os.path.exists(plots[key]):
                # Add section title
                title_cell = ws_graphs[cell]
                title_cell.value = title
                title_cell.font = Font(bold=True, size=14)
                
                # Add the image
                img = Image(plots[key])
                
                # Resize to fit nicely
                img.width = 600
                img.height = 360
                
                # Position the image (one row below the title)
                img_row = int(cell[1:]) + 1
                img_cell = f"{cell[0]}{img_row}"
                
                ws_graphs.add_image(img, img_cell)
                
                print(f"Added {key} plot at {img_cell}")
        
        # Adjust column widths
        ws_graphs.column_dimensions['A'].width = 15
        ws_graphs.column_dimensions['J'].width = 15
        
        # Save the workbook
        wb.save(output_file)
        print(f"Excel report with graphs saved to: {output_file}")
        
        # Clean up temporary plot files
        import shutil
        if os.path.exists(plot_dir):
            shutil.rmtree(plot_dir)
            print("Cleaned up temporary plot files")
    
    def _generate_basic_excel(self, output_file):
        """Generate basic Excel report (without graphs)"""
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Analysis Duration (s)',
                    'Average Frequency (Hz)',
                    'Average Jitter (ms)',
                    'Average Bandwidth (Mbps)',
                    'Total Messages',
                    'Total Data (MB)',
                    'Average CPU Usage (%)',
                    'Average Memory Usage (%)',
                    'Max Temperature (°C)',
                    'Performance Rating',
                    'Stability Rating'
                ],
                'Value': [
                    self.data['summary']['analysis_duration'],
                    f"{self.data['lidar_metrics']['average_hz']:.2f}",
                    f"{self.data['lidar_metrics']['average_jitter_ms']:.2f}",
                    f"{self.data['lidar_metrics']['average_bandwidth_mbps']:.2f}",
                    self.data['lidar_metrics']['total_messages'],
                    f"{self.data['lidar_metrics']['total_mb']:.2f}",
                    f"{self.data['system_resources']['average_cpu_percent']:.1f}",
                    f"{self.data['system_resources']['average_memory_percent']:.1f}",
                    f"{self.data['system_resources'].get('max_temperature_c', 'N/A')}",
                    self.data['analysis']['performance_rating'],
                    self.data['analysis']['stability_rating']
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format Summary sheet
            worksheet = writer.sheets['Summary']
            for idx, col in enumerate(df_summary.columns):
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = 30
            
            # Detailed metrics sheet
            metrics_dict = {
                'Frequency (Hz)': [
                    self.data['lidar_metrics']['average_hz'],
                    self.data['lidar_metrics']['min_hz'],
                    self.data['lidar_metrics']['max_hz']
                ],
                'Jitter (ms)': [
                    self.data['lidar_metrics']['average_jitter_ms'],
                    self.data['lidar_metrics']['min_jitter_ms'],
                    self.data['lidar_metrics']['max_jitter_ms']
                ],
                'Bandwidth (Mbps)': [
                    self.data['lidar_metrics']['average_bandwidth_mbps'],
                    self.data['lidar_metrics']['min_bandwidth_mbps'],
                    self.data['lidar_metrics']['max_bandwidth_mbps']
                ]
            }
            
            df_metrics = pd.DataFrame(metrics_dict, index=['Average', 'Minimum', 'Maximum'])
            df_metrics.to_excel(writer, sheet_name='Metrics Statistics')
            
            # Recommendations sheet
            if 'recommendations' in self.data['analysis']:
                rec_data = {
                    'Recommendation': self.data['analysis']['recommendations']
                }
                df_rec = pd.DataFrame(rec_data)
                df_rec.to_excel(writer, sheet_name='Recommendations', index=False)


def main():
    parser = argparse.ArgumentParser(description='Generate enhanced Excel report with embedded graphs')
    parser.add_argument('--json-file', type=str, required=True,
                       help='Path to benchmark JSON report')
    parser.add_argument('--viz-data', type=str, 
                       default='/tmp/lidar_benchmark/visualization_data.json',
                       help='Path to visualization data JSON')
    parser.add_argument('--output', type=str,
                       default='/tmp/lidar_benchmark_report_enhanced.xlsx',
                       help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Check if files exist
    if not os.path.exists(args.json_file):
        print(f"Error: JSON file not found: {args.json_file}")
        return
    
    if args.viz_data and not os.path.exists(args.viz_data):
        print(f"Warning: Visualization data not found: {args.viz_data}")
        print("Generating report without graphs...")
    
    # Generate report
    generator = EnhancedExcelReport(args.json_file, args.viz_data)
    generator.generate_excel_with_graphs(args.output)


if __name__ == '__main__':
    main()