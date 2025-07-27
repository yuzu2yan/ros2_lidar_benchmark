#!/usr/bin/env python3

import json
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
import numpy as np
import os
import argparse
from openpyxl import load_workbook
from openpyxl.drawing.image import Image


class ExcelReportWithGraphs:
    def __init__(self, json_file, visualization_data_file=None):
        """Initialize report generator with data files"""
        self.json_file = json_file
        self.visualization_data_file = visualization_data_file
        
        # Load benchmark data
        with open(json_file, 'r') as f:
            self.data = json.load(f)
        
        # Load visualization data if available
        self.viz_data = None
        if visualization_data_file and os.path.exists(visualization_data_file):
            with open(visualization_data_file, 'r') as f:
                self.viz_data = json.load(f)
        
        # Ensure pandas is available
        try:
            import pandas as pd
        except ImportError:
            print("Error: pandas is required. Install with: pip install pandas openpyxl")
            exit(1)
    
    def create_time_series_plots(self, output_dir):
        """Create time series plots for all metrics"""
        if not self.viz_data or len(self.viz_data.get('timestamps', [])) == 0:
            print("No visualization data available for plotting")
            return {}
        
        # Create output directory
        os.makedirs(output_dir, exist_ok=True)
        
        # Plot configuration
        plt.style.use('seaborn-v0_8-darkgrid')
        fig_size = (10, 6)
        
        plots = {}
        
        # 1. Frequency (Hz) plot
        if 'hz' in self.viz_data and len(self.viz_data['hz']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'], self.viz_data['hz'], 'b-', linewidth=2)
            plt.title('LiDAR Publishing Frequency Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Frequency (Hz)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            # Add average line
            avg_hz = np.mean(self.viz_data['hz'])
            plt.axhline(y=avg_hz, color='r', linestyle='--', label=f'Average: {avg_hz:.1f} Hz')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'frequency_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['frequency'] = plot_file
        
        # 2. Jitter plot
        if 'jitter' in self.viz_data and len(self.viz_data['jitter']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'], self.viz_data['jitter'], 'r-', linewidth=2)
            plt.title('Message Jitter Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Jitter (ms)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            # Add average line
            avg_jitter = np.mean(self.viz_data['jitter'])
            plt.axhline(y=avg_jitter, color='g', linestyle='--', label=f'Average: {avg_jitter:.1f} ms')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'jitter_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['jitter'] = plot_file
        
        # 3. Bandwidth plot
        if 'bandwidth' in self.viz_data and len(self.viz_data['bandwidth']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'], self.viz_data['bandwidth'], 'g-', linewidth=2)
            plt.title('Network Bandwidth Usage Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Bandwidth (Mbps)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            avg_bw = np.mean(self.viz_data['bandwidth'])
            plt.axhline(y=avg_bw, color='orange', linestyle='--', label=f'Average: {avg_bw:.1f} Mbps')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'bandwidth_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['bandwidth'] = plot_file
        
        # 4. Throughput plot
        if 'throughput' in self.viz_data and len(self.viz_data['throughput']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'], self.viz_data['throughput'], 'm-', linewidth=2)
            plt.title('Point Cloud Throughput Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Throughput (K points/sec)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            avg_tp = np.mean(self.viz_data['throughput'])
            plt.axhline(y=avg_tp, color='blue', linestyle='--', label=f'Average: {avg_tp:.1f} K pts/s')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'throughput_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['throughput'] = plot_file
        
        # 5. CPU Usage plot
        if 'cpu' in self.viz_data and len(self.viz_data['cpu']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'][:len(self.viz_data['cpu'])], 
                    self.viz_data['cpu'], 'orange', linewidth=2)
            plt.title('CPU Usage Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('CPU Usage (%)', fontsize=12)
            plt.ylim(0, 105)
            plt.grid(True, alpha=0.3)
            
            avg_cpu = np.mean(self.viz_data['cpu'])
            plt.axhline(y=avg_cpu, color='red', linestyle='--', label=f'Average: {avg_cpu:.1f}%')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'cpu_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['cpu'] = plot_file
        
        # 6. Memory Usage plot
        if 'memory' in self.viz_data and len(self.viz_data['memory']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'][:len(self.viz_data['memory'])], 
                    self.viz_data['memory'], 'c-', linewidth=2)
            plt.title('Memory Usage Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Memory Usage (%)', fontsize=12)
            plt.ylim(0, 105)
            plt.grid(True, alpha=0.3)
            
            avg_mem = np.mean(self.viz_data['memory'])
            plt.axhline(y=avg_mem, color='blue', linestyle='--', label=f'Average: {avg_mem:.1f}%')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'memory_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['memory'] = plot_file
        
        # 7. Temperature plot
        if 'temperature' in self.viz_data and len(self.viz_data['temperature']) > 0:
            plt.figure(figsize=fig_size)
            plt.plot(self.viz_data['timestamps'][:len(self.viz_data['temperature'])], 
                    self.viz_data['temperature'], 'r-', linewidth=2)
            plt.title('CPU Temperature Over Time', fontsize=14, fontweight='bold')
            plt.xlabel('Time (seconds)', fontsize=12)
            plt.ylabel('Temperature (°C)', fontsize=12)
            plt.grid(True, alpha=0.3)
            
            avg_temp = np.mean(self.viz_data['temperature'])
            max_temp = np.max(self.viz_data['temperature'])
            plt.axhline(y=avg_temp, color='orange', linestyle='--', label=f'Average: {avg_temp:.1f}°C')
            plt.axhline(y=max_temp, color='red', linestyle=':', label=f'Max: {max_temp:.1f}°C')
            plt.legend()
            
            plot_file = os.path.join(output_dir, 'temperature_plot.png')
            plt.savefig(plot_file, dpi=150, bbox_inches='tight')
            plt.close()
            plots['temperature'] = plot_file
        
        return plots
    
    def generate_excel_with_graphs(self, output_file):
        """Generate Excel report with embedded graphs"""
        # First, generate the basic Excel report
        self._generate_basic_excel(output_file)
        
        # Create plots
        plot_dir = os.path.dirname(output_file)
        plots = self.create_time_series_plots(plot_dir)
        
        if not plots:
            print("No plots generated, skipping graph integration")
            return
        
        # Load the workbook and add graphs
        wb = load_workbook(output_file)
        
        # Create a new sheet for graphs
        if 'Graphs' in wb.sheetnames:
            del wb['Graphs']
        
        ws_graphs = wb.create_sheet('Graphs')
        ws_graphs.sheet_properties.tabColor = "FF0000"
        
        # Add title
        ws_graphs['A1'] = 'Performance Metrics Time Series Graphs'
        ws_graphs['A1'].font = ws_graphs['A1'].font.copy(bold=True, size=16)
        
        # Add plots to the sheet
        row = 3
        col = 'A'
        
        plot_info = [
            ('frequency', 'Publishing Frequency'),
            ('jitter', 'Message Jitter'),
            ('bandwidth', 'Network Bandwidth'),
            ('throughput', 'Point Cloud Throughput'),
            ('cpu', 'CPU Usage'),
            ('memory', 'Memory Usage'),
            ('temperature', 'CPU Temperature')
        ]
        
        for key, title in plot_info:
            if key in plots:
                # Add title
                ws_graphs[f'{col}{row}'] = title
                ws_graphs[f'{col}{row}'].font = ws_graphs[f'{col}{row}'].font.copy(bold=True, size=12)
                
                # Add image
                img = Image(plots[key])
                img.width = 600
                img.height = 400
                ws_graphs.add_image(img, f'{col}{row + 1}')
                
                # Move to next position
                row += 23  # Space for image and gap
        
        # Save the workbook
        wb.save(output_file)
        print(f"Excel report with graphs saved to: {output_file}")
        
        # Clean up temporary plot files
        for plot_file in plots.values():
            if os.path.exists(plot_file):
                os.remove(plot_file)
    
    def _generate_basic_excel(self, output_file):
        """Generate basic Excel report (without graphs)"""
        # This is a simplified version of the original excel_report_generator
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': [
                    'Analysis Duration (s)',
                    'Average Frequency (Hz)',
                    'Average Jitter (ms)',
                    'Average Bandwidth (Mbps)',
                    'Total Messages',
                    'Total Data (MB)',
                    'Average CPU Usage (%)',
                    'Average Memory Usage (%)',
                    'Performance Rating',
                    'Stability Rating'
                ],
                'Value': [
                    self.data['summary']['analysis_duration'],
                    self.data['lidar_metrics']['average_hz'],
                    self.data['lidar_metrics']['average_jitter_ms'],
                    self.data['lidar_metrics']['average_bandwidth_mbps'],
                    self.data['lidar_metrics']['total_messages'],
                    self.data['lidar_metrics']['total_mb'],
                    self.data['system_resources']['average_cpu_percent'],
                    self.data['system_resources']['average_memory_percent'],
                    self.data['analysis']['performance_rating'],
                    self.data['analysis']['stability_rating']
                ]
            }
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Detailed metrics sheet
            df_metrics = pd.DataFrame([self.data['lidar_metrics']])
            df_metrics.to_excel(writer, sheet_name='Detailed Metrics', index=False)
            
            # System resources sheet
            df_system = pd.DataFrame([self.data['system_resources']])
            df_system.to_excel(writer, sheet_name='System Resources', index=False)
            
            # Raw data sheet
            df_raw = pd.DataFrame([self.data])
            df_raw.to_excel(writer, sheet_name='Raw Data', index=False)


def main():
    parser = argparse.ArgumentParser(description='Generate Excel report with time series graphs')
    parser.add_argument('--json-file', type=str, required=True,
                       help='Path to benchmark JSON report')
    parser.add_argument('--viz-data', type=str, 
                       default='/tmp/lidar_benchmark/visualization_data.json',
                       help='Path to visualization data JSON')
    parser.add_argument('--output', type=str,
                       default='/tmp/lidar_benchmark_report_with_graphs.xlsx',
                       help='Output Excel file path')
    
    args = parser.parse_args()
    
    # Generate report
    generator = ExcelReportWithGraphs(args.json_file, args.viz_data)
    generator.generate_excel_with_graphs(args.output)


if __name__ == '__main__':
    main()