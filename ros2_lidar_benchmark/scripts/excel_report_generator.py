#!/usr/bin/env python3

import json
import os
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import argparse


class ExcelReportGenerator:
    def __init__(self, json_report_path, output_path=None, visualization_data_path=None):
        self.json_report_path = json_report_path
        self.visualization_data_path = visualization_data_path
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f'/tmp/lidar_benchmark_report_{timestamp}.xlsx'
        
        self.output_path = output_path
        self.data = None
        self.viz_data = None
        self.wb = None
        
    def load_json_report(self):
        """Load JSON report data"""
        with open(self.json_report_path, 'r') as f:
            self.data = json.load(f)
        
        # Load visualization data if available
        if self.visualization_data_path and os.path.exists(self.visualization_data_path):
            try:
                with open(self.visualization_data_path, 'r') as f:
                    self.viz_data = json.load(f)
            except Exception as e:
                print(f"Warning: Could not load visualization data: {e}")
                self.viz_data = None
    
    def create_excel_report(self):
        """Create comprehensive Excel report"""
        self.wb = Workbook()
        
        # Remove default sheet
        self.wb.remove(self.wb.active)
        
        # Create sheets
        self._create_summary_sheet()
        self._create_metrics_sheet()
        self._create_system_resources_sheet()
        self._create_top_processes_sheet()
        self._create_raw_data_sheet()
        
        # Save workbook
        self.wb.save(self.output_path)
        print(f"Excel report saved to: {self.output_path}")
    
    def _create_summary_sheet(self):
        """Create summary sheet with overview information"""
        ws = self.wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "ROS 2 LiDAR Benchmark Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Test information
        row = 3
        info_data = [
            ("Test Start Time", self.data.get('start_time', 'N/A')),
            ("Test End Time", self.data.get('end_time', 'N/A')),
            ("Duration (seconds)", self.data.get('duration_seconds', 'N/A')),
            ("Performance Rating", self.data['summary'].get('performance_rating', 'N/A')),
            ("Stability Rating", self.data['summary'].get('stability_rating', 'N/A'))
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Performance metrics summary
        row += 1
        ws[f'A{row}'] = "Performance Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        headers = ['Metric', 'Average', 'Min', 'Max', 'Std Dev']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        if self.data.get('metrics'):
            metrics = self.data['metrics']
            
            # Frequency data
            row += 1
            ws[f'A{row}'] = "Frequency (Hz)"
            ws[f'B{row}'] = f"{metrics['frequency']['average_hz']:.2f}"
            ws[f'C{row}'] = f"{metrics['frequency']['min_hz']:.2f}"
            ws[f'D{row}'] = f"{metrics['frequency']['max_hz']:.2f}"
            ws[f'E{row}'] = f"{metrics['frequency']['std_hz']:.2f}"
            
            # Jitter data
            row += 1
            ws[f'A{row}'] = "Jitter (ms)"
            ws[f'B{row}'] = f"{metrics['jitter']['average_ms']:.2f}"
            ws[f'C{row}'] = f"{metrics['jitter']['min_ms']:.2f}"
            ws[f'D{row}'] = f"{metrics['jitter']['max_ms']:.2f}"
            ws[f'E{row}'] = f"{metrics['jitter']['std_ms']:.2f}"
            
            # Bandwidth data
            row += 1
            ws[f'A{row}'] = "Bandwidth (Mbps)"
            ws[f'B{row}'] = f"{metrics['bandwidth']['average_mbps']:.2f}"
            ws[f'C{row}'] = f"{metrics['bandwidth']['min_mbps']:.2f}"
            ws[f'D{row}'] = f"{metrics['bandwidth']['max_mbps']:.2f}"
            
            # Throughput data if available
            if 'throughput' in metrics:
                row += 1
                ws[f'A{row}'] = "Throughput (K pts/s)"
                ws[f'B{row}'] = f"{metrics['throughput']['average_kpoints_per_sec']:.2f}"
                ws[f'C{row}'] = "-"
                ws[f'D{row}'] = f"{metrics['throughput']['max_kpoints_per_sec']:.2f}"
        
        # Recommendations
        row += 2
        ws[f'A{row}'] = "Recommendations"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:E{row}')
        
        row += 1
        for rec in self.data['summary'].get('recommendations', []):
            ws[f'A{row}'] = f"• {rec}"
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_metrics_sheet(self):
        """Create detailed metrics sheet with charts"""
        ws = self.wb.create_sheet("Metrics Details")
        
        # Title
        ws['A1'] = "Detailed Performance Metrics"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not self.data.get('metrics'):
            ws['A3'] = "No metrics data available"
            return
        
        metrics = self.data['metrics']
        
        # Create data table
        row = 3
        ws['A3'] = "Metric"
        ws['B3'] = "Value"
        ws['C3'] = "Unit"
        
        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws[cell].font = Font(color="FFFFFF", bold=True)
        
        metrics_data = [
            ("Average Frequency", metrics['frequency']['average_hz'], "Hz"),
            ("Min Frequency", metrics['frequency']['min_hz'], "Hz"),
            ("Max Frequency", metrics['frequency']['max_hz'], "Hz"),
            ("Frequency Std Dev", metrics['frequency']['std_hz'], "Hz"),
            ("", "", ""),
            ("Average Jitter", metrics['jitter']['average_ms'], "ms"),
            ("Min Jitter", metrics['jitter']['min_ms'], "ms"),
            ("Max Jitter", metrics['jitter']['max_ms'], "ms"),
            ("Jitter Std Dev", metrics['jitter']['std_ms'], "ms"),
            ("", "", ""),
            ("Average Bandwidth", metrics['bandwidth']['average_mbps'], "Mbps"),
            ("Min Bandwidth", metrics['bandwidth']['min_mbps'], "Mbps"),
            ("Max Bandwidth", metrics['bandwidth']['max_mbps'], "Mbps"),
            ("", "", ""),
            ("Avg Messages/sec", metrics.get('throughput', {}).get('average_messages_per_sec', 0), "msg/s"),
            ("Avg MB/sec", metrics.get('throughput', {}).get('average_mbytes_per_sec', 0), "MB/s"),
            ("Avg K Points/sec", metrics.get('throughput', {}).get('average_kpoints_per_sec', 0), "Kpts/s"),
            ("Max K Points/sec", metrics.get('throughput', {}).get('max_kpoints_per_sec', 0), "Kpts/s"),
            ("Avg Points/Message", metrics.get('throughput', {}).get('average_points_per_message', 0), "points"),
            ("", "", ""),
            ("Total Samples", metrics.get('samples', 0), "count")
        ]
        
        for i, (metric, value, unit) in enumerate(metrics_data, 4):
            ws[f'A{i}'] = metric
            if isinstance(value, (int, float)):
                ws[f'B{i}'] = f"{value:.2f}" if unit != "count" else str(value)
            else:
                ws[f'B{i}'] = str(value)
            ws[f'C{i}'] = unit
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 20
    
    def _create_system_resources_sheet(self):
        """Create system resources sheet"""
        ws = self.wb.create_sheet("System Resources")
        
        # Title
        ws['A1'] = "System Resource Usage"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not self.data.get('system_stats'):
            ws['A3'] = "No system resource data available"
            return
        
        stats = self.data['system_stats']
        
        # Create data table
        row = 3
        headers = ['Resource', 'Average', 'Min', 'Max']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # CPU data
        row += 1
        ws[f'A{row}'] = "CPU Usage (%)"
        ws[f'B{row}'] = f"{stats['cpu']['average_percent']:.1f}"
        ws[f'C{row}'] = f"{stats['cpu']['min_percent']:.1f}"
        ws[f'D{row}'] = f"{stats['cpu']['max_percent']:.1f}"
        
        # Memory data
        row += 1
        ws[f'A{row}'] = "Memory Usage (%)"
        ws[f'B{row}'] = f"{stats['memory']['average_percent']:.1f}"
        ws[f'C{row}'] = f"{stats['memory']['min_percent']:.1f}"
        ws[f'D{row}'] = f"{stats['memory']['max_percent']:.1f}"
        
        # Temperature data if available
        if 'temperature' in stats:
            row += 1
            ws[f'A{row}'] = "CPU Temperature (°C)"
            ws[f'B{row}'] = f"{stats['temperature']['average_celsius']:.1f}"
            ws[f'C{row}'] = f"{stats['temperature']['min_celsius']:.1f}"
            ws[f'D{row}'] = f"{stats['temperature']['max_celsius']:.1f}"
        
        # Jetson-specific temperatures if available
        if 'jetson_temperatures' in stats:
            row += 2
            ws[f'A{row}'] = "Jetson Temperature Zones"
            ws[f'A{row}'].font = Font(bold=True)
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            for zone, temps in stats['jetson_temperatures'].items():
                ws[f'A{row}'] = f"{zone.upper()} (°C)"
                ws[f'B{row}'] = f"{temps['average_celsius']:.1f}"
                ws[f'C{row}'] = f"{temps['min_celsius']:.1f}"
                ws[f'D{row}'] = f"{temps['max_celsius']:.1f}"
                row += 1
        
        # Additional info
        row += 1
        ws[f'A{row}'] = "Total Samples:"
        ws[f'B{row}'] = stats.get('samples', 0)
        ws[f'A{row}'].font = Font(bold=True)
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
    
    def _create_top_processes_sheet(self):
        """Create top processes sheet with memory and CPU usage"""
        ws = self.wb.create_sheet("Top Processes")
        
        # Title
        ws['A1'] = "Top 20 Processes Resource Usage"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not self.viz_data or 'top_processes_memory' not in self.viz_data:
            ws['A3'] = "No top processes data available"
            return
        
        # Calculate statistics for each process
        processes_memory = self.viz_data.get('top_processes_memory', [])
        processes_cpu = self.viz_data.get('top_processes_cpu', [])
        timestamps = self.viz_data.get('timestamps', [])
        
        if not processes_memory or len(processes_memory) == 0:
            ws['A3'] = "No top processes data available"
            return
        
        # Create summary table
        row = 3
        headers = ['Process #', 'Process Name', 'Memory Avg (MB)', 'Memory Min (MB)', 'Memory Max (MB)', 
                   'CPU Avg (%)', 'CPU Min (%)', 'CPU Max (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Get process names
        process_names = self.viz_data.get('top_processes_names', [])
        
        # Calculate statistics for each process
        for i in range(min(20, len(processes_memory))):
            row += 1
            mem_data = processes_memory[i] if i < len(processes_memory) else []
            cpu_data = processes_cpu[i] if i < len(processes_cpu) else []
            
            # Filter out NaN values
            import math
            mem_valid = [x for x in mem_data if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))]
            cpu_valid = [x for x in cpu_data if isinstance(x, (int, float)) and not (isinstance(x, float) and math.isnan(x))]
            
            # Get process name
            proc_name = process_names[i] if i < len(process_names) and process_names[i] else f'Process {i+1}'
            
            ws[f'A{row}'] = f"Process {i+1}"
            ws[f'B{row}'] = proc_name
            
            if len(mem_valid) > 0:
                ws[f'C{row}'] = f"{sum(mem_valid) / len(mem_valid):.2f}"
                ws[f'D{row}'] = f"{min(mem_valid):.2f}"
                ws[f'E{row}'] = f"{max(mem_valid):.2f}"
            else:
                ws[f'C{row}'] = "N/A"
                ws[f'D{row}'] = "N/A"
                ws[f'E{row}'] = "N/A"
            
            if len(cpu_valid) > 0:
                ws[f'F{row}'] = f"{sum(cpu_valid) / len(cpu_valid):.2f}"
                ws[f'G{row}'] = f"{min(cpu_valid):.2f}"
                ws[f'H{row}'] = f"{max(cpu_valid):.2f}"
            else:
                ws[f'F{row}'] = "N/A"
                ws[f'G{row}'] = "N/A"
                ws[f'H{row}'] = "N/A"
        
        # Add note about data source
        row += 2
        ws[f'A{row}'] = "Note: Processes are ranked by memory usage at the time of monitoring."
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:H{row}')
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 12  # Process #
        ws.column_dimensions['B'].width = 30  # Process Name (wider for names)
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws.column_dimensions[col].width = 18
    
    def _create_raw_data_sheet(self):
        """Create raw data sheet with all JSON data"""
        ws = self.wb.create_sheet("Raw Data")
        
        # Title
        ws['A1'] = "Complete JSON Report Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Convert JSON to formatted string
        json_str = json.dumps(self.data, indent=2)
        lines = json_str.split('\n')
        
        # Write JSON data
        for i, line in enumerate(lines, 3):
            ws[f'A{i}'] = line
        
        # Set column width
        ws.column_dimensions['A'].width = 100
    
    def generate(self):
        """Main method to generate the report"""
        print(f"Loading JSON report from: {self.json_report_path}")
        self.load_json_report()
        
        print("Generating Excel report...")
        self.create_excel_report()
        
        return self.output_path


def main():
    parser = argparse.ArgumentParser(description='Generate Excel report from JSON benchmark data')
    parser.add_argument('--json-file', 
                       default='/tmp/lidar_benchmark_report.json',
                       help='Path to JSON report file')
    parser.add_argument('--output-file',
                       help='Output Excel file path (default: auto-generated)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.json_file):
        print(f"Error: JSON file not found: {args.json_file}")
        return 1
    
    generator = ExcelReportGenerator(args.json_file, args.output_file)
    output_path = generator.generate()
    
    print(f"\nExcel report successfully generated: {output_path}")
    return 0


if __name__ == '__main__':
    exit(main())